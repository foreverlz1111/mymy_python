import os
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart
import time
import pprint

# from openpyxl.utils import get_column_letter
# from datetime import datetime

############################
# github.com/foreverlz1111 #
############################
__author__ = "foreverlz1111"
__license__ = "GPL3.0"


########
# 变量 #
########
files_name = {
    "answer": "answer.xlsx",
    "students": "students.xlsx",
    "question": "question.xlsx",
}
dir_name = {"students_score_out": "out", "students_answer": "student"}


########
# 常量 #
########

__uname = os.uname()
__current_cwd = os.getcwd()

__denote = (
    "*启动目录自动新建"
    + dir_name["students_score_out"]
    + "文件夹存放学生成绩.xlsx文件、"
    + dir_name["students_answer"]
    + "文件夹存放学生答案.xlsx文件\n*默认在启动文件夹内读取answer.xlsx文件作为答案,读取students.xlsx文件作为花名册\n"
)
__borderline = "************************************************************"
dic_name = {
    "students_score_out": "成绩输出目录",
    "students_answer": "作答输入目录",
    "keyboard_exit": "\n你选择退出，请重新启动程序",
    "add_student": "\n将文件放入指定目录后(文件夹不会被检测到)，按下[回车键]以继续",
    "answer_notfound": "没找到答案文件，请确保其在启动目录并命名为" + files_name["answer"],
    "students_notfound": "没找到花名册文件，请确保其在启动目录并命名为" + files_name["students"],
    "answer_found": __current_cwd + " 目录已找到文件：" + files_name["answer"],
    "students_found": __current_cwd + " 目录已找到文件：" + files_name["students"],
    "choice_notfound": "没找到对应命令，请重新输入。",
    "processing": "处理中...",
    "processing_current_file":"当前文件：{}",
    "score_title": ["学号", "姓名", "班级", "成绩", "题目总分"],
    "missing_title": ["学号", "姓名", "班级", "未提交"],
    "office_title":["题号", "正确", "总分"],
    "yes": "是",
    "save_error": "处理失败！",
    "save_success": "处理成功，查看文件 ",
    "saved_path": __current_cwd + "/" + dir_name["students_score_out"] + "/",
}

__menu_name = {
    "menu": "\n选择菜单：\n",
    "choice1": "输入1：重新查找文件",
    "choice2": "输入2：输出成绩文件",
    "choice3": "输入3：输出未提交作业的学生名单文件 ",
    "choice4": "输入4：修改计算机等级二级office - excel题目",
    "choice_exit": "输入q：退出本程序",
}


def print_menu():
    for x in __menu_name:
        print(__menu_name[x])
    print(__borderline)


def catch_choice(c):
    # 版本兼容性，暂不升级为match
    if c == "1":
        check_files()
        print_menu()
        return True
    elif c == "2":
        output_score()
        print_menu()
        return True
    elif c == "3":
        output_missing()
        print_menu()
        return True
    elif c == "4":
        office_excel()
        print_menu()
        return True
    elif c == "q":
        return False
    else:
        print(dic_name["choice_notfound"])
        print_menu()
        return True


def menu():
    print_menu()
    userexit = True
    while userexit:
        try:
            tmp = input()
            userexit = catch_choice(tmp)
            time.sleep(1)
        except KeyboardInterrupt:
            print(dic_name["keyboard_exit"])
            exit()


def office_excel():
    # check one at once
    print(dic_name["processing"])
    with os.scandir(dir_name["students_answer"]) as scan:
        for entry in scan:
            if entry.is_file() and entry.name.endswith(".xlsx"):
                # do
                filename=(dir_name["students_answer"] + "/" + entry.name)
                single_dataonly = load_workbook(
                    filename=filename,
                    data_only=True,
                    )
                single = load_workbook(filename=filename)
                single_dataonly_sheet1 = single_dataonly["Sheet1"]
                single_dataonly_sheet2 = single_dataonly["Sheet2"]
                single_sheet1 = single["Sheet1"]
                single_sheet2 = single["Sheet2"]
                single_sheet3 = single["Sheet3"]
                single_sheet4 = single["Sheet4"]
                print(dic_name["processing_current_file"].format(filename))
                wb = Workbook()
                ws = wb.active
                ws.append(dic_name["office_title"])
                tb = []
                tmp = []
                try:
                    tmp.clear()
                    tmp.append(1)
                    if office_excel_q1(single_sheet4):
                        tmp.append("是")
                        tmp.append(12.5)
                    else:
                        tmp.append("否")
                        tmp.append(0)
                    ws.append(tmp.copy())
                    tb.append(tmp.copy())
                except Exception as e:
                    print(e)
                try:
                    tmp.clear()
                    tmp.append(2)
                    if office_excel_q2(single_sheet4):
                        tmp.append("是")
                        tmp.append(12.5)
                    else:
                        tmp.append("否")
                        tmp.append(0)
                    ws.append(tmp.copy())
                    tb.append(tmp.copy())
                except Exception as e:
                    print(e)
                try:
                    tmp.clear()
                    tmp.append(3)
                    if office_excel_q3(single_sheet1):
                        tmp.append("是")
                        tmp.append(12.5)
                    else:
                        tmp.append("否")
                        tmp.append(0)
                    ws.append(tmp.copy())
                    tb.append(tmp.copy())
                except Exception as e:
                    print(e)
                try:
                    tmp.clear()
                    tmp.append(4)
                    if office_excel_q4(single_sheet1):
                        tmp.append("是")
                        tmp.append(12.5)
                    else:
                        tmp.append("否")
                        tmp.append(0)
                    ws.append(tmp.copy())
                    tb.append(tmp.copy())
                except Exception as e:
                    print(e)
                try:
                    tmp.clear()
                    tmp.append(5)
                    if office_excel_q5(single_sheet1):
                        tmp.append("是")
                        tmp.append(12.5)
                    else:
                        tmp.append("否")
                        tmp.append(0)
                    ws.append(tmp.copy())
                    tb.append(tmp.copy())
                except Exception as e:
                    print(e)
                try:
                    tmp.clear()
                    tmp.append(6)
                    if office_excel_q6(single_dataonly_sheet1,single_sheet1):
                        tmp.append("是")
                        tmp.append(12.5)
                    else:
                        tmp.append("否")
                        tmp.append(0)
                    ws.append(tmp.copy())
                    tb.append(tmp.copy())
                except Exception as e:
                    print(e)
                try:
                    tmp *= 0
                    tmp.append(7)
                    if office_excel_q7(single_dataonly_sheet2):
                        tmp.append("是")
                        tmp.append(12.5)
                    else:
                        tmp.append("否")
                        tmp.append(0)
                    ws.append(tmp.copy())
                    tb.append(tmp.copy())
                except Exception as e:
                    print(e)
                try:
                    tmp *= 0
                    tmp.append(8)
                    if office_excel_q8(single_sheet3):
                        tmp.append("是")
                        tmp.append(12.5)
                    else:
                        tmp.append("否")
                        tmp.append(0)
                    ws.append(tmp.copy())
                    tb.append(tmp.copy())
                except Exception as e:
                    print(e)
                
                tmp.clear()
                sum = 0
                tmp.append("总分")
                for v in tb:
                    sum += v[2]
                tmp.append(sum)
                ws.append(tmp)
                try:
                    wb.save(dir_name["students_score_out"] + "/" + output_xlsx_name(entry.name[:len(entry.name)-5]+"成绩"))
                    print(dic_name["save_success"], dic_name["saved_path"] + output_xlsx_name(entry.name[:len(entry.name)-5]+"成绩"))
                except OSError:
                    print(dic_name["save_error"])   
        scan.close()
        print(__borderline)


def office_excel_q1(s):
    # trying find validation
    block = "A1"
    vtype = "textLength"
    voperator = "lessThanOrEqual"
    verror = "只能录入5位数字或文本"
    verrorstyle = "warning"
    for x in s.data_validations.dataValidation:
        # row text by print(x)
        if x.sqref == block:
            if x.type == vtype and x.operator == voperator:
                if x.errorStyle == verrorstyle:
                    if x.error == verror:
                        print("q1 —— pass")
                        return True
    print("q1 —— error")
    return False

def office_excel_q2(s):
    # get time formula
    block = "C1"
    block_answer = '=MROUND(B1,"0:15")'
    c1 = s[block]
    if c1.value == block_answer:
        print("q2 —— pass")
        return True
    print("q2 —— error")
    return False

def office_excel_q3(s):
    # code upgrade formula only
    added_pos = 3
    added_str = "0"
    formula = "=REPLACE({},{},0,{})"
    for x in s.iter_rows(min_row=3, min_col=2, max_col=3):
        if not x[1].value == formula.format(
            x[0].coordinate, str(added_pos), str(added_str)
        ):
            print("q3 —— error")
            return False
    print("q3 —— pass")
    return True

def office_excel_q4(s):
    # formulae included below,suggest the first:
    # DATEDIF(?,TODAY(),"y") or DATEDIF(?,NOW(),"y")
    # INT(YEARFRAC(?,TODAY(),1)) or INT(YEARFRAC(?,TODAY())) or INT(YEARFRAC(?,NOW()),1) or INT(YEARFRAC(?,NOW()))
    # YEAR(TODAY()-?+2)-1900 or YEAR(NOW()-?+2)-1900
    formulae = [
        '=DATEDIF({},TODAY(),"y")',
        '=DATEDIF({},NOW(),"y")',
        '=DATEDIF(E3,NOW(),"y")' "=INT(YEARFRAC({},TODAY(),1))",
        "=INT(YEARFRAC({},TODAY()))",
        "=INT(YEARFRAC({},NOW()),1)",
        "=INT(YEARFRAC({},NOW()))",
        "=YEAR(TODAY()-{}+2)-1900",
        "=YEAR(NOW()-{}+2)-1900",
    ]
    # age , worked year
    for x in s.iter_rows(min_row=3, min_col=5, max_col=6):
        found = False
        for f in formulae:
            if not x[1].value == f.format(x[0].coordinate):
                pass
            else:
                found = True
                break
        if not found:
            print("q4 —— error")
            return found
    for x in s.iter_rows(min_row=3, min_col=7, max_col=8):
        found = False
        for f in formulae:
            if not x[1].value == f.format(x[0].coordinate):
                pass
            else:
                found = True
                break
        if not found:
            print("q4 —— error")
            return found
    print("q4 —— pass")
    return True

def office_excel_q5(s):
    block1 = "N3"
    block1_answer = '=COUNTIF(D$3:D$66,"男")'
    block2 = "N4"
    block2_answer = '=COUNTIF(I$3:I$66,"高级工程师")'
    block3 = "N5"
    block3_answer = '=COUNTIF(H$3:H$66,">=10")'
    n3 = s[block1]
    n4 = s[block2]
    n5 = s[block3]
    n = [n3, n4, n5]
    a = [block1_answer, block2_answer, block3_answer]
    for x,v in enumerate(n):
        if not v.value == a[x]:
            print("q5 —— error")
            return False
    print("q5 —— pass")
    return True

def office_excel_q6(s1,s2):
    # ""IF"" function only
    ift = "\"TRUE\""
    ifn = "\"FALSE\""
    formula = "=IF(AND({}>20,{}=\"工程师\"),{},{})"
    for x in s2.iter_rows(min_row=3, min_col=8, max_col=11):
        if not x[3].value == formula.format(s1.cell(row=x[0].row,column=x[0].column).coordinate,x[1].coordinate,ift,ifn):
            print("q6 —— error")
            return Flase
    print("q6 —— pass")
    return True


def office_excel_q7(s):
    # check value
    choices = ["男",30,10,"助工"]
    for x in s.iter_rows(min_row=3, min_col=4, max_col=9):
        if not (x[0].value == choices[0] and x[2].value >= choices[1] and  x[4].value >= choices[2] and x[5].value==choices[3]):
            print("q7 —— error")
            return False
    print("q7 —— pass")
    return True
    
def office_excel_q8(s):
    # chart sheet
    selected_answer = "职称"
    x_axis_answer = "职称"
    y_axis_answer = "职称"
    p = s._pivots[0]
    p.cache.refreshOnLoad = True
    if not s.cell(row=p.location.firstDataRow,column=p.location.firstDataCol).value == selected_answer:
        print("q8 —— error")
        return False
    for t in s._charts:
        # print("print(t.x_axis.title.body)")
        # print(t.x_axis.title.body)
        # print(__borderline)
        # print("print(t.x_axis.title.extLst)")
        # print(t.x_axis.title.extLst)
        # print(__borderline)
        # print("print(t.x_axis.title.graphicalProperties)")
        # print(t.x_axis.title.graphicalProperties)
        # print(__borderline)
        # print("print(t.x_axis.title.layout)")
        # print(t.x_axis.title.layout)
        # print(__borderline)
        # print("print(t.x_axis.title.overlay)")
        # print(t.x_axis.title.overlay)
        # print(__borderline)
        # print("print(t.x_axis.title.spPr)")
        # print(t.x_axis.title.spPr)
        # print(__borderline)
        if not t.x_axis.title.text.rich.p[0].r[0].t == x_axis_answer:
            print("q8 —— error")
            return False
        if not t.y_axis.title.text.rich.p[0].r[0].t == y_axis_answer:
            print("q8 —— error")
            return False
        # print("print(t.x_axis.title.text)")
        # print(t.x_axis.title.text)
        # print(__borderline)
        # print("print(t.x_axis.title.text.rich.p[0].r[0].t)")
        # print(t.x_axis.title.text.rich.p[0].r[0].t)
        # print(__borderline)
        # print("print(t.x_axis.title.tx)")
        # print(t.x_axis.title.tx)
        # print(__borderline)
        # print("print(t.x_axis.title.txPr")
        # print(t.x_axis.title.txPr)
        # print(__borderline)
        # print(t.y_axis.title)
        # print(__borderline)
        # for v in t.ser:
            # print("print(v.bubbleSize)")
            # print(v.bubbleSize)
            # print("print(v.dLbls)")
            # print(v.dLbls)
            # print("print(v.cat)")
            # print(v.cat)
            # print("print(v.dPt)")
            # print(v.dPt)
            # print("print(v.data_points)")
            # print(v.data_points)
            # print("print(v.errBars)")
            # print(v.errBars)
            # print("print(v.extLst)")
            # print(v.extLst)
            # print("v.graphicalPropertie")
            # print(v.graphicalProperties)
            # print("print(v.identifiers)")
            # print(v.identifiers)
            # print("print(v.idx)")
            # print(v.idx)
            # print("print(v.invertIfNegative)")
            # print(v.invertIfNegative)
            # print("print(v.labels)")
            # print(v.labels)
            # print("print(v.marker)")
            # print(v.marker)
            # print("print(v.order)")
            # print(v.order)
            # print("print(v.pictureOptions)")
            # print(v.pictureOptions)
            # print("print(v.title)")
            # print(v.title)
            # print("print(v.spPr)")
            # print(v.spPr)
            # print("print(v.trendline)")
            # print(v.trendline)
            # print("print(v.tx)")
            # print(v.tx.strRef)
            # print("print(v.val)")
            # print(v.val)
            # print("print(v.xVal)")
            # print(v.xVal)
    print("q8 —— pass")
    return True

def output_missing():
    # only check file if missing
    print(dic_name["processing"])
    students = load_workbook(filename=files_name["students"])

    # if windows :
    # students_sheet1 = students["Sheet"]  # 输出的成绩花名册
    # if linux :
    # students_sheet1 = students["工作表1"]  # 输出的成绩花名册

    if sys.platform.startswith("linux"):
        students_sheet1 = students["工作表1"]
    elif sys.platform.startswith("win32"):
        students_sheet1 = students["Sheet"]

    students_table = get_table(students_sheet1)
    wb = Workbook()
    ws = wb.active
    ws.append(dic_name["missing_title"])
    for x in students_table:
        try:
            p = students_answer_filename(x[0])
            if not os.path.exists(p):
                x.append(dic_name["yes"])
                ws.append(x)
        except:
            continue
    try:
        wb.save(dir_name["students_score_out"] + "/" + output_xlsx_name("未提交"))
        print(
            dic_name["save_success"], dic_name["saved_path"] + output_xlsx_name("未提交")
        )
    except OSError:
        print(dic_name["save_error"])
    print(__borderline)


def output_score():
    # check file for each
    print(dic_name["processing"])
    answer_sheet1 = load_and_sheet1(files_name["answer"])
    answer_table = get_table(answer_sheet1)
    students = load_workbook(filename=files_name["students"])

    # if windows :
    # students_sheet1 = students["Sheet"]  # 输出的成绩花名册
    # if linux :
    # students_sheet1 = students["工作表1"]  # 输出的成绩花名册

    if sys.platform.startswith("linux"):
        students_sheet1 = students["工作表1"]
    elif sys.platform.startswith("win32"):
        students_sheet1 = students["Sheet"]

    students_table = get_table(students_sheet1)
    wb = Workbook()
    ws = wb.active
    score_total = get_total(answer_table)
    ws.append(dic_name["score_title"])
    for x in students_table:
        try:
            single = load_workbook(filename=students_answer_filename(x[0]))
            single_sheet1 = single["Sheet"]
            single_table = get_table(single_sheet1)
        except FileNotFoundError:
            pass
        score = 0
        for a in single_table:
            for b in answer_table:
                if a[0] == b[0]:
                    if a[1] == b[1]:
                        score += b[2]
                else:
                    continue
        x.append(score)
        x.append(score_total)
        ws.append(x)
    try:
        wb.save(dir_name["students_score_out"] + "/" + output_xlsx_name("成绩"))
        print(dic_name["save_success"], dic_name["saved_path"] + output_xlsx_name("成绩"))
    except OSError:
        print(dic_name["save_error"])
    print(__borderline)


def students_answer_filename(x):
    return dir_name["students_answer"] + "/" + str(x) + ".xlsx"


def output_xlsx_name(s):
    mytime = time.strftime("%Y%m%d%H%M%S", time.localtime())
    filename = s + str(mytime) + ".xlsx"
    return filename


def get_total(t):
    score = 0
    for c in t:
        score += c[2]
    return score


def load_and_sheet1(f):
    wb = load_workbook(filename=f)

    # if windows :
    # sheet1 = wb["Sheet"]
    # if linux :
    # sheet1 = wb["工作表1"]

    if sys.platform.startswith("linux"):
        sheet1 = wb["工作表1"]
    elif sys.platform.startswith("win32"):
        sheet1 = wb["Sheet"]

    return sheet1


def get_table(sheet):
    cells = []
    table = []
    p = False
    for x in sheet.iter_rows(min_row=2):
        cells = []
        for cell in x:
            if cell.value is None:
                break
            cells.append(cell.value)
        if cells != []:
            table.append(cells)
    return table


def get_sysinfo():
    if sys.platform.startswith("linux"):
        print(__uname.sysname, __uname.release, __uname.machine)
        # time.sleep(1)
        # os.chdir("/home")
    elif sys.platform.startswith("win32"):
        print("Windows")
        # os.chdir("c:\windows")


def get_varinfo():
    print("当前启动目录", __current_cwd)
    create_dir()


def create_dir():
    dir_error = "位置异常"
    dir_exist = "已存在，跳过新建"
    dir_mk = "已自动新建文件夹"
    for x in dir_name:
        try:
            p = __current_cwd + "/" + dir_name[x]
            if not os.path.exists(p):
                try:
                    os.mkdir(p)
                    print(dic_name[x], p, dir_mk)
                    continue
                except FileNotFoundError:
                    print(dir_error)
            print(dic_name[x], p, dir_exist)
        except:
            continue


def start():
    print(__borderline)
    get_sysinfo()
    print(__denote)
    get_varinfo()
    print(__borderline)
    print(dic_name["add_student"])
    try:
        tmp = input()
        # print(tmp)
        # time.sleep(1)
    except KeyboardInterrupt:
        print(dic_name["keyboard_exit"])
        exit()


def check_files():
    scan_student()  # 学生文件夹
    scan_answer()  # 答案文件
    scan_student_list()  # 花名册文件
    print(__borderline)


def scan_student():
    files_count = 0
    student_dir = __current_cwd + "/" + dir_name["students_answer"]
    with os.scandir(student_dir) as scan:
        for entry in scan:
            if entry.is_file() and entry.name.endswith(".xlsx"):
                files_count += 1
    scan.close()
    print(student_dir, "目录已找到xlsx文件数量：", files_count)


def scan_answer():
    try:
        fp = open(files_name["answer"])
    except FileNotFoundError:
        print(dic_name["answer_notfound"])
    else:
        print(dic_name["answer_found"])
        fp.close()


def scan_student_list():
    try:
        fp = open(files_name["students"])
    except FileNotFoundError:
        print(dic_name["students_notfound"])
    else:
        print(dic_name["students_found"])
        fp.close()


if __name__ == "__main__":
    start()
    check_files()
    menu()
