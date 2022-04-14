from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook
import os
import random
import time
def test1():
    wb = load_workbook(filename = "test.xlsx")
    # 加载后 用 wb.save("test_new.xlsx")保存
    
    sheet_ranges = wb["工作表1"]
    count = 0
    # sheet_ranges['']遍历列号
    for x in sheet_ranges['B']:
        count += 1
        print(x.value)
        if count == 10:
            break
    print(count,"\n")
    
    # iter_rows(起始行，计算列数，计算行数)
    for row in sheet_ranges.iter_rows(min_row = 1,max_col=3,max_row=2):
        for cell in row:
            print(cell.coordinate,"格子位置")
            print(cell.value)
    # print(sheet_ranges["A1"].value)

    #sheet_ranges.insert_rows(7) 第七行为新增，即截断前面6行
    #sheet_ranges.delete_cols(1,2)从第几列开始删除，删除几列
    #sheet_ranges.delete_rows(7) 第七行为删除，后面不会补上

    # 完全遍历
    for row in sheet_ranges:
        break
        for value in row:
            print(value.value)
   # wb.save("test_new.xlsx")

    
    wb = Workbook()
    ws = wb.active
    
    data = [
        ['Apples', 10000, 5000, 8000, 6000],
        ['Pears',   2000, 3000, 4000, 5000],
        ['Bananas', 6000, 6000, 6500, 6000],
        ['Oranges',  500,  300,  200,  700],
    ]

    # add column headings. NB. these must be strings
    ws.append(["Fruit", "2011", "2012", "2013", "2014"])
    for row in data:
        ws.append(row)
        
    tab = Table(displayName="Table1", ref="B1:E1")

    # Add a default style with striped rows and banded columns
    # style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    # tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save("table.xlsx")
    '''
    Table must be added using ws.add_table() method to avoid duplicate names.
Using this method ensures table name is unque through out defined names and all other table name. 
    '''
    scan = os.scandir()
    files_count = 0
    for entry in scan:
        if entry.is_file() and entry.name.endswith(""):
            print(entry.name)
            files_count += 1
    print(files_count)
    scan.close()

def test2():
    wb = load_workbook(filename = "students.xlsx")
    sheet_ranges = wb["工作表1"]
    count = 0
    # iter_rows(起始行，计算列数，计算行数)
    for row in sheet_ranges.iter_rows(min_row = 2,max_col=3,max_row=3):
        for cell in row:
            count += 1
            print(cell.coordinate,"格子位置")
            print(cell.value)
            print(count)
def test3():
    #生成一份答案文档
    wb = Workbook()
    ws1 = wb.active
    title = ["编号","答案","分值"]
    datas = []
    alpha = ["A","B","C","D","E","F"]
    count = 1
    ws1.append(title)
    while count < 21:
        datas.append(count)
        datas.append(random.choice(alpha))
        datas.append(random.randint(1,5))
        count += 1
        ws1.append(datas)
        print(datas)
        datas = []
    wb.save("test3.xlsx")

def test4():
    #生成学生答案
    stu = load_workbook(filename = "students.xlsx")
    title = ["编号","答案"]
    datas = []
    alpha = ["A","B","C","D","E","F"]
    sheet_ranges = stu["工作表1"]
    for x in sheet_ranges.iter_rows(min_row = 2,max_col=1):
         for cell in x:
            if cell.value is None:
                break
            wb = Workbook()
            ws1 = wb.active
            ws1.append(title)
            count = 1
            number = cell.value
            while count < 21:
                datas.append(count)
                datas.append(random.choice(alpha))
                count += 1
                ws1.append(datas)
                datas = []
                wb.save("student/"+str(number)+".xlsx"
)
if __name__ == "__main__":
    print("begin!")
            
    


    
    
